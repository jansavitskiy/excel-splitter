import openpyxl
import os
import re
from openpyxl.drawing.image import Image
from pathlib import Path
from openpyxl.utils import get_column_letter
from .utils import ensure_directory_exists, adjust_column_widths, copy_cell_styles, sanitize_filename

class ExcelSplitter:
    def __init__(self, messages, language):
        self.messages = messages
        self.language = language
        self.image_path = None

    def set_image(self, image_path):
        """Устанавливает путь к изображению для вставки"""
        if image_path and os.path.exists(image_path):
            self.image_path = image_path
            return True
        return False

    def process_all_cards(self, input_file, output_folder, card_marker, width_percent):
        """Разделяет Excel файл на отдельные карточки"""
        if not os.path.exists(input_file):
            print(self.messages['file_not_found'])
            return False

        try:
            wb = openpyxl.load_workbook(input_file)
            ws = wb.active
            card_count = 0
            card_start = 1
            
            print(self.messages['processing'])
            
            for i, row in enumerate(ws.iter_rows(), 1):
                if any(cell.value and card_marker in str(cell.value) for cell in row):
                    if card_count > 0:
                        self._save_card(ws, card_start, i-1, output_folder, 
                                      card_count, card_marker, width_percent/100)
                    card_start = i
                    card_count += 1
            
            if card_count > 0:
                self._save_card(ws, card_start, ws.max_row, output_folder,
                              card_count, card_marker, width_percent/100)
            
            print(self.messages['total_cards'].format(count=card_count))
            print(self.messages['done'].format(path=output_folder))
            return True
            
        except Exception as e:
            error_msg = f"{self.messages['error']}: {str(e)}"
            print(error_msg)
            return False

    def _save_card(self, src_sheet, start_row, end_row, output_folder, 
                  card_num, card_marker, width_factor):
        """Сохраняет карточку с изображением (отступ не более 2 строк)"""
        try:
            os.makedirs(output_folder, exist_ok=True)
            
            wb = openpyxl.Workbook()
            ws = wb.active
            
            # Копирование данных
            new_last_row = 1
            for row in range(start_row, end_row + 1):
                for col in range(1, src_sheet.max_column + 1):
                    src_cell = src_sheet.cell(row=row, column=col)
                    dst_cell = ws.cell(row=new_last_row, column=col)
                    dst_cell.value = src_cell.value
                    copy_cell_styles(src_cell, dst_cell)
                new_last_row += 1
            
            # Добавление изображения
            if self.image_path:
                self._add_image_to_sheet(ws, new_last_row)
            
            # Настройка ширины столбцов
            self._copy_sheet_properties(src_sheet, ws, start_row, end_row)
            adjust_column_widths(ws, width_factor)
            
            # Сохранение файла
            filename = self._generate_filename(ws, card_num, card_marker)
            filepath = os.path.join(output_folder, filename)
            wb.save(filepath)
            print(self.messages['card_saved'].format(filename=filename))
            
        except Exception as e:
            print(f"{self.messages['error']}: Ошибка при сохранении карточки")
            print(f"Подробности: {str(e)}")

    def _add_image_to_sheet(self, ws, last_row):
        """Добавляет изображение в конец листа (макс 2 строки отступа)"""
        try:
            img = Image(self.image_path)
            
            # Фиксируем оригинальные размеры
            img.width = 300  # Ширина в пикселях
            img.height = 150  # Высота в пикселях
            
            # Позиционирование (максимум 2 строки отступа)
            anchor_col = 'A'
            anchor_row = last_row + min(2, 3)  # Не более 2 строк отступа
            
            # Добавление изображения
            ws.add_image(img, f"{anchor_col}{anchor_row}")
            
            # Настройка высоты строки
            ws.row_dimensions[anchor_row].height = img.height * 0.75
            
        except Exception as e:
            print(f"{self.messages['warning']}: Не удалось добавить изображение")
            print(f"Подробности: {str(e)}")

    def _copy_sheet_properties(self, src_sheet, dst_sheet, start_row, end_row):
        """Копирует свойства листа"""
        for col in range(1, src_sheet.max_column + 1):
            col_letter = get_column_letter(col)
            if col_letter in src_sheet.column_dimensions:
                dst_sheet.column_dimensions[col_letter].width = \
                    src_sheet.column_dimensions[col_letter].width
        
        for row in range(start_row, end_row + 1):
            if row in src_sheet.row_dimensions:
                dst_sheet.row_dimensions[row-start_row+1].height = \
                    src_sheet.row_dimensions[row].height
        
        for merged_range in src_sheet.merged_cells.ranges:
            if merged_range.min_row >= start_row and merged_range.max_row <= end_row:
                dst_sheet.merge_cells(
                    start_row=merged_range.min_row-start_row+1,
                    end_row=merged_range.max_row-start_row+1,
                    start_column=merged_range.min_col,
                    end_column=merged_range.max_col
                )

    def _generate_filename(self, sheet, card_num, card_marker):
        """Генерирует безопасное имя файла"""
        card_number = f"{card_num:03d}"
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value and card_marker in str(cell.value):
                    try:
                        card_number = str(cell.value).split(card_marker)[-1].strip()
                        card_number = re.sub(r'[<>:"/\\|?*]', '', card_number)
                        card_number = card_number.strip()
                        if not card_number:
                            card_number = f"{card_num:03d}"
                        break
                    except:
                        pass
        
        prefix = "Card" if self.language == 'E' else "Карточка"
        filename = f"{prefix}_{card_num:03d}_{card_number}.xlsx"
        return sanitize_filename(filename)