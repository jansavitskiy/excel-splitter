import openpyxl
import os
from pathlib import Path
from .utils import copy_cell_styles, validate_output_path


class ExcelMerger:
    def __init__(self, messages, language):
        self.messages = messages
        self.language = language


    def merge_files(self, input_folder, output_file):
        """Объединяет все Excel файлы из папки в один"""
        if not os.path.exists(input_folder):
            print(self.messages['folder_not_found'])
            return False

        try:
            output_file = validate_output_path(output_file)
            merged_wb = openpyxl.Workbook()
            merged_ws = merged_wb.active
            merged_ws.title = "Merged" if self.language == 'E' else "Объединенные данные"

            file_count = 0
            row_counter = 1

            for filename in sorted(os.listdir(input_folder)):
                if filename.endswith(('.xlsx', '.xls')):
                    filepath = os.path.join(input_folder, filename)
                    try:
                        wb = openpyxl.load_workbook(filepath)
                        ws = wb.active

                        # Копируем данные
                        for row in ws.iter_rows():
                            for col_idx, cell in enumerate(row, 1):
                                merged_ws.cell(row=row_counter, column=col_idx, value=cell.value)
                                copy_cell_styles(cell, merged_ws.cell(row=row_counter, column=col_idx))
                            row_counter += 1

                        file_count += 1
                    except Exception as e:
                        print(f"{self.messages['error_reading']} {filename}: {str(e)}")

            if file_count > 0:
                Path(output_file).parent.mkdir(parents=True, exist_ok=True)
                merged_wb.save(output_file)
                print(self.messages['merge_success'].format(count=file_count, path=output_file))
                return True
            else:
                print(self.messages['no_files_found'])
                return False

        except Exception as e:
            print(f"{self.messages['error']}: {str(e)}")
            return False