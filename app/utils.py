import os
import re
from pathlib import Path
from copy import copy
from openpyxl.utils import get_column_letter

def ensure_directory_exists(path):
    """Создаёт папку с обработкой ошибок для Windows"""
    try:
        Path(path).mkdir(parents=True, exist_ok=True)
    except Exception as e:
        raise Exception(f"Не удалось создать папку {path}: {str(e)}")


def sanitize_filename(filename):
    """Очищает имя файла от недопустимых символов"""
    # Заменяем запрещённые символы на подчёркивание
    filename = re.sub(r'[<>:"/\\|?*]', '_', filename)
    # Удаляем начальные и конечные пробелы
    filename = filename.strip()
    # Убедимся, что имя не пустое
    if not filename:
        filename = "unnamed_file"
    return filename

def ensure_directory_exists(path):
    Path(path).mkdir(parents=True, exist_ok=True)


def adjust_column_widths(sheet, width_factor):
    """Изменяет ширину столбцов, не затрагивая изображения"""
    for col in range(1, sheet.max_column + 1):
        col_letter = get_column_letter(col)
        if col_letter in sheet.column_dimensions:
            original_width = sheet.column_dimensions[col_letter].width
            sheet.column_dimensions[col_letter].width = original_width * width_factor


def copy_cell_styles(src_cell, dst_cell):
    if src_cell.has_style:
        dst_cell.font = copy(src_cell.font)
        dst_cell.border = copy(src_cell.border)
        dst_cell.fill = copy(src_cell.fill)
        dst_cell.number_format = src_cell.number_format
        dst_cell.alignment = copy(src_cell.alignment)


def validate_output_path(path):
    path = Path(path)
    if not path.suffix:
        path = path.with_suffix('.xlsx')
    return str(path)