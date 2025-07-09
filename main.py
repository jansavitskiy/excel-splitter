import os
from openpyxl import load_workbook
try:
    from openpyxl.drawing.image import Image
    from PIL import Image as PILImage  # Дополнительная проверка Pillow
    print("Pillow успешно импортирован")
except ImportError as e:
    raise ImportError("Не удалось импортировать Pillow. Установите его командой: pip install pillow") from e

def insert_image_to_excel(input_file, output_file, image_path):
    try:
        print(f"\nОбработка файла: {os.path.basename(input_file)}")
        
        # Проверка существования файлов
        if not os.path.exists(input_file):
            raise FileNotFoundError(f"Входной файл не найден: {input_file}")
        if not os.path.exists(image_path):
            raise FileNotFoundError(f"Изображение не найдено: {image_path}")

        # Загрузка изображения с проверкой
        try:
            img = Image(image_path)
            print(f"Изображение {image_path} успешно загружено")
        except Exception as img_error:
            raise ValueError(f"Ошибка загрузки изображения: {str(img_error)}") from img_error

        # Работа с Excel
        wb = load_workbook(input_file)
        ws = wb.active
        
        # Позиция для вставки
        img_row = ws.max_row + 2
        img_col = 1
        
        # Вставка изображения
        ws.add_image(img, f'A{img_row}')
        
        # Сохранение
        wb.save(output_file)
        print(f"Успешно сохранено в {output_file}")
        return True
    
    except Exception as e:
        print(f"Ошибка при обработке файла {input_file}: {str(e)}")
        return False

if __name__ == '__main__':
    input_folder = "fax"
    output_folder = "Карточки_с_подписью"
    image_path = "img.png"

    # Создаем папку для результатов
    os.makedirs(output_folder, exist_ok=True)

    # Обрабатываем файлы
    success_count = 0
    for filename in os.listdir(input_folder):
        if filename.endswith('.xlsx'):
            input_path = os.path.join(input_folder, filename)
            output_path = os.path.join(output_folder, filename)
            
            if insert_image_to_excel(input_path, output_path, image_path):
                success_count += 1

    print(f"\nОбработка завершена. Успешно обработано файлов: {success_count}")